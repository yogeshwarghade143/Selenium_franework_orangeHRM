import time

import openpyxl
# import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By

driver = webdriver.Chrome()
driver.implicitly_wait(10)
driver.get("https://opensource-demo.orangehrmlive.com/")
driver.maximize_window()

path = "E:\\software testing\\lectures\selenium_framework\\testData\\Book1.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook["Sheet5"]
total_row=sheet.max_row
print(total_row)
time.sleep(3)

driver.implicitly_wait(10)


for i in range(2, total_row+1 ):

    username = sheet.cell(row=i, column=1).value
    password = sheet.cell(row=i, column=2).value
    print(f"username : {username}, password : {password}")


    time.sleep(3)
    email = driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[1]/div/div[2]/div[2]/form/div[1]/div/div[2]/input')
                                        # '/html/body/div/div[1]/div/div[1]/div/div[2]/div[2]/form/div[1]/div/div[2]/input'
    email.send_keys(username)
    time.sleep(3)
    driver.find_element(By.NAME, 'password').send_keys(password)
    time.sleep(3)

    driver.find_element(By.XPATH, '/html/body/div/div[1]/div/div[1]/div/div[2]/div[2]/form/div[3]/button').click()
    time.sleep(3)
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[1]/header/div[1]/div[2]/ul/li/span/i').click()
    time.sleep(3)
    driver.find_element(By.XPATH, '/html/body/div/div[1]/div[1]/header/div[1]/div[2]/ul/li/ul/li[4]/a').click()
    time.sleep(3)













